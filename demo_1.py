from bs4 import BeautifulSoup
import time
import xlwt
import re
from lxml import etree
import requests
from openpyxl.styles import Alignment
from requests import ConnectionError
from openpyxl import Workbook


class QcWy:
    def __init__(self):
        self.url = "https://search.51job.com/list/090200%252c060000,000000,0000,00,9,04,%257B%257D,2," \
              "{}.html?lang=c&postchannel=0000&workyear=01&cotype=99&degreefrom=04&jobterm=99&companysize=99" \
                   "&ord_field=0&dibiaoid=0&line=&welfare= "
        # https://search.51job.com/list/090200%252c060000,000000,0000,00,9,04,%257B%257D,2,2.html?lang=c&postchannel=0000&workyear=01&cotype=99&degreefrom=04&jobterm=99&companysize=99&ord_field=0&dibiaoid=0&line=&welfare=
        # https://search.51job.com/list/090200%252c060000,000000,0000,00,9,04,%257B%257D,2,3.html?lang=c&postchannel=0000&workyear=01&cotype=99&degreefrom=04&jobterm=99&companysize=99&ord_field=0&dibiaoid=0&line=&welfare=
        # https://search.51job.com/list/090200%252c060000,000000,0000,00,9,04,%257B%257D,2,4.html?lang=c&postchannel=0000&workyear=01&cotype=99&degreefrom=04&jobterm=99&companysize=99&ord_field=0&dibiaoid=0&line=&welfare=
        # 定义请求头
        self.headers = {
            "cache-control": "no-cache",
            "postman-token": "72a56deb-825e-3ac3-dd61-4f77c4cbb4d8",
            "Host": "search.51job.com",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) "
                          "Chrome/70.0.3538.67 Safari/537.36",
        }
        self.row = 2
        self.list_url = []  # 存放每个岗位url
        self.position = []  # 存放职位信息
        self.company = []   # 存放公司信息
        self.gw_name = []  # 岗位
        self.gz = []  # 工资
        self.gs_url = []  # 公司url
        self.gs_name = []  # 公司名称
        self.position_url = []  # 查看所有职位url
        self.text = []  # 地址时间等等信息
        self.qt = []  # 其他信息
        self.work = []  # 上班地点

    def creat_execl(self):
        wordbook = Workbook()  # 创建工作簿
        word = wordbook.active
        wordbook.worksheets[0].title = "前程无忧信息"
        rows = ["工作岗位", "公司", "公司-url", "查看所有职位-url", "工资", "地点", "工作年限", "学历要求", "需求人数",
                "发布日期", "其他", "联系方式", "职位信息", "公司信息"]
        word.append(rows)
        self.get_content(sheet1=word)
        return wordbook

    def get_page_url(self, number):
        global text
        try:
            response = requests.get(url=self.url.format(number), headers=self.headers)
            if response.status_code == 200:
                text = response.text
                print("***" * 12 + "成功获取到第" + str(number) + "个页面的url！" + "**" * 12)
            else:
                print("***" * 12 + "没有获取到第" + str(number) + "个页面的url！" + "**" * 12)
        except ConnectionError as e:
            print(e.args)
        # print(response.encoding)  # 输出该页面的编码
        # print(response.text)
        # response.encoding = "utf-8"   # 修改页面的编码
        # print(response.encoding)
        # print(response.text)
        title = re.compile('"job_href":"(.*?)"', re.S)  # 获取每个岗位的url
        title_ = re.findall(title, text)
        # print(title_)
        for value_url in title_:
            if "\\" in value_url:
                new_url = value_url.replace("\\", "")
                self.list_url.append(new_url)
                # print(new_url)

    def get_content(self, sheet1):
        global response1
        for url in self.list_url:
            try:
                response1 = requests.get(url=url, headers=self.headers)
                # print(res.encoding)  # 查看原网页的编码
                response1.encoding = "gbk"  # 打印text时，显示乱码则设置编码
                # print(res.text)
            except:
                print("连接失败")
            html = etree.HTML(response1.text)
            text_ = html.xpath('//div[@class="cn"]')
            for values in text_:
                self.gw_name.append(values.xpath('./h1/text()'))
                self.gz.append(values.xpath('./strong/text()'))
                self.gs_url.append(values.xpath('./p[@class="cname"]/a[1]/@href'))
                self.gs_name.append(values.xpath('./p[@class="cname"]/a[1]/text()'))
                self.position_url.append(values.xpath('./p[@class="cname"]/a[2]/@href'))
                self.text.append(values.xpath('./p[@class="msg ltype"]/text()'))
                self.qt.append(values.xpath('./div[1]/div[1]/span/text()'))

            soup = BeautifulSoup(response1.text, "html.parser")
            position = soup.find_all("div", class_="bmsg job_msg inbox")
            for zw_xx in position:
                self.position.append(zw_xx.get_text())
                # print(zw_xx.get_text())

            work = soup.find_all("div", class_="bmsg inbox")
            for work_ in work:
                self.work.append(work_.get_text())
                # print(work_.get_text())

            company = soup.find_all("div", class_="tmsg inbox")  # 定位公司信息
            for gs_xx in company:
                self.company.append(gs_xx.get_text())
                # print(gs_xx.get_text())
        # print(self.gw_name)
        # print(self.gs_name)
        # print(self.gs_url)
        # print(self.gz)
        # print(self.position_url)
        # print(self.text)
        # print(self.qt)
        for value in zip(self.gw_name, self.gs_name, self.gs_url, self.position_url, self.gz, self.text, self.qt):
            print(value)
            # cell使用 openpyxl 的首行、首列 是 （1,1）而不是（0,0），如果坐标输入含有小于1的值，提示 ：Row or column values must be at least 1，即最小值为1.
            sheet1.cell(self.row, 1, value[0][0])
            sheet1.cell(self.row, 2, value[1][0])
            sheet1.cell(self.row, 3, value[2][0])
            sheet1.cell(self.row, 4, value[3][0])
            try:
                sheet1.cell(self.row, 5, value[4][0])
            except:
                pass
            sheet1.cell(self.row, 6, value[5][0])
            sheet1.cell(self.row, 7, value[5][1])
            sheet1.cell(self.row, 8, value[5][2])
            sheet1.cell(self.row, 9, value[5][3])
            sheet1.cell(self.row, 10, value[5][4])
            str_qt = ""
            for value_1 in value[6]:
                str_qt += (value_1 + ' ')
                sheet1.cell(self.row, 11, str_qt)
            self.row += 1

        row = 2
        for value_2 in zip(self.work, self.position, self.company):
            sheet1.cell(row, 12, value_2[0])
            sheet1.cell(row, 13, value_2[1])
            sheet1.cell(row, 14, value_2[2])
            row += 1
        self.set_excel(sheet1)

    def set_excel(self, word_set):
        """
        设置Excel表格的格式
        :param word_set: 表格对象
        :return:
        """
        tuple_1 = ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N")
        list_dyg = ["A{}", "B{}", "C{}", "D{}", "E{}", "F{}", "G{}", "H{}", "I{}", "J{}", "K{}", "L{}", "M{}", "N{}"]
        # 设置每行高度为30
        for value_h in range(1, len(self.gs_name) + 2):
            word_set.row_dimensions[value_h].height = 40

        # 设置每列宽度为20
        for tuple_value in tuple_1:
            word_set.column_dimensions[tuple_value].width = 40

        # 设置每个单元格内容居中
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for value_dyg in list_dyg:
            for value_1 in range(1, len(self.gs_name) + 2):
                word_set[value_dyg.format(value_1)].alignment = align

    def text_save(self):
        pass

    def mysql_save(self):
        pass


if __name__ == '__main__':
    QC = QcWy()
    print("**" * 10 + "开始爬取前程无忧的岗位信息" + "**" * 10)
    start_page = int(input("请输入开始爬取的页面："))
    end_page = int(input("请输入爬取结束的页面："))
    for num in range(start_page, end_page + 1):
        QC.get_page_url(num)
    QC.creat_execl().save("./前程无忧.xlsx")   # 已经调用改函数
print('\n')
print("**" * 10 + "结束爬取前程无忧的岗位信息" + "**" * 10)
